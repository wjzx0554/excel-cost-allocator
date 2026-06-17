from pathlib import Path

from openpyxl import Workbook, load_workbook

from excel_cost_allocator.allocator import (
    AllocationConfig,
    AllocationScheme,
    BatchAllocationConfig,
    FilterRule,
    allocate_workbook,
    allocate_workbook_batch,
    create_sample_workbook,
    get_headers,
    get_unique_values,
    preview_filter_matches,
    preview_workbook_batch,
)
from excel_cost_allocator.templates import import_scheme_template, serialize_scheme_template


def test_allocate_with_filter_and_multiple_targets(tmp_path: Path):
    input_path = tmp_path / "input.xlsx"
    output_path = tmp_path / "output.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "sheet1"
    ws.append(["说明", None, None, None, None])
    ws.append(["车间", "材料", "人工", "共耗料", "水电费"])
    ws.append(["生产车间", 100, 0, 10, 20])
    ws.append(["销售配货部", 100, 100, 30, 40])
    ws.append(["生产车间", 300, 100, 60, 40])
    wb.save(input_path)

    result = allocate_workbook(
        AllocationConfig(
            input_path=str(input_path),
            output_path=str(output_path),
            sheet_name="sheet1",
            header_row=2,
            base_columns=[2, 3],
            allocation_columns=[4, 5],
            filter_column=1,
            excluded_values={"销售配货部"},
        )
    )

    assert result.total_rows == 3
    assert result.participating_rows == 2
    assert result.excluded_rows == 1

    out = load_workbook(output_path, data_only=True)
    ws = out["sheet1"]

    assert ws["D3"].value == 20
    assert ws["D4"].value == 0
    assert ws["D5"].value == 80
    assert ws["E3"].value == 20
    assert ws["E4"].value == 0
    assert ws["E5"].value == 80

    assert "分摊明细" in out.sheetnames
    detail = out["分摊明细"]
    assert detail["B3"].value == 3
    assert detail["D3"].value == 2
    assert detail["F3"].value == 1


def test_headers_keep_trailing_blank_title_columns(tmp_path: Path):
    path = tmp_path / "headers.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "sheet1"
    ws["A1"] = "表头1"
    ws["C1"] = "表头3"
    ws["D2"] = 123
    ws["F5"] = 456
    wb.save(path)

    headers = get_headers(str(path), "sheet1", 1)
    assert len(headers) >= 6
    assert headers[1].header == ""
    assert headers[2].header == "表头3"
    assert headers[5].header == ""


def test_unique_values_reads_filter_column(tmp_path: Path):
    path = tmp_path / "values.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "sheet1"
    ws.append(["车间", "金额"])
    ws.append(["生产车间", 10])
    ws.append(["销售配货部", 20])
    ws.append(["生产车间", 30])
    ws.append([None, 40])
    wb.save(path)

    assert get_unique_values(str(path), "sheet1", 1, 1) == ["生产车间", "销售配货部", ""]


def test_multi_rule_filter_logic_and_regex(tmp_path: Path):
    path = tmp_path / "rules.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "sheet1"
    ws.append(["车间", "备注", "金额"])
    ws.append(["生产车间", "临时", 10])
    ws.append(["销售配货部", "正式", 20])
    ws.append(["售后服务部", "临时", 30])
    wb.save(path)

    rules = [
        FilterRule(column=1, operator="equals", value="销售配货部"),
        FilterRule(column=2, operator="regex", value="^临时$"),
    ]

    count_or, samples_or = preview_filter_matches(str(path), "sheet1", 1, rules, "OR")
    count_and, samples_and = preview_filter_matches(str(path), "sheet1", 1, rules, "AND")

    assert count_or == 3
    assert count_and == 0
    assert len(samples_or) == 3
    assert samples_and == []


def test_batch_allocation_with_manual_amount_and_different_filters(tmp_path: Path):
    input_path = tmp_path / "batch.xlsx"
    output_path = tmp_path / "batch_out.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "sheet1"
    ws.append(["车间", "类别", "材料", "人工", "共耗料", "运费"])
    ws.append(["一车间", "成品", 100, 100, 0, 0])
    ws.append(["二车间", "成品", 300, 100, 0, 0])
    ws.append(["销售配货部", "内部", 100, 100, 0, 0])
    wb.save(input_path)

    config = BatchAllocationConfig(
        input_path=str(input_path),
        output_path=str(output_path),
        sheet_name="sheet1",
        header_row=1,
        schemes=[
            AllocationScheme(
                name="共耗料分摊",
                amount_source="manual",
                manual_amount=1000,
                allocation_column=5,
                base_columns=[3, 4],
                filter_rules=[FilterRule(column=1, operator="equals", value="销售配货部")],
            ),
            AllocationScheme(
                name="运费分摊",
                amount_source="manual",
                manual_amount=300,
                allocation_column=6,
                base_columns=[3],
                filter_rules=[FilterRule(column=2, operator="regex", value="^内部$")],
            ),
        ],
    )

    preview = preview_workbook_batch(config)
    assert len(preview.scheme_results) == 2
    assert preview.scheme_results[0].target_total == 1000
    assert preview.scheme_results[0].participating_rows == 2
    assert preview.scheme_results[1].base_total == 400

    result = allocate_workbook_batch(config)
    assert result.total_rows == 3
    assert result.scheme_results[0].distributed_total == 1000
    assert result.scheme_results[1].distributed_total == 300

    out = load_workbook(output_path, data_only=True)
    ws = out["sheet1"]
    assert ws["E2"].value == 333.33
    assert ws["E3"].value == 666.67
    assert ws["E4"].value == 0
    assert ws["F2"].value == 75
    assert ws["F3"].value == 225
    assert ws["F4"].value == 0
    assert "分摊汇总" in out.sheetnames
    assert "分摊明细" not in out.sheetnames
    assert "明细_01_共耗料分摊" in out.sheetnames
    assert "明细_02_运费分摊" in out.sheetnames
    summary = out["分摊汇总"]
    assert summary["A1"].value == "费用分摊汇总"
    assert summary["B6"].value == "明细_01_共耗料分摊"
    assert summary["B7"].value == "明细_02_运费分摊"
    first_detail = out["明细_01_共耗料分摊"]
    second_detail = out["明细_02_运费分摊"]
    assert first_detail["A1"].value == "方案明细：共耗料分摊"
    assert second_detail["A1"].value == "方案明细：运费分摊"


def test_batch_allocation_with_source_column_amount(tmp_path: Path):
    input_path = tmp_path / "source_column.xlsx"
    output_path = tmp_path / "source_column_out.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "sheet1"
    ws.append(["车间", "材料", "人工", "待分摊运费", "运费分摊"])
    ws.append(["一车间", 100, 100, 1000, 0])
    ws.append(["二车间", 300, 100, 2000, 0])
    wb.save(input_path)

    config = BatchAllocationConfig(
        input_path=str(input_path),
        output_path=str(output_path),
        sheet_name="sheet1",
        header_row=1,
        schemes=[
            AllocationScheme(
                name="运费分摊",
                amount_source="column_total",
                amount_column=4,
                allocation_column=5,
                base_columns=[2, 3],
                filter_rules=[],
            )
        ],
    )

    result = allocate_workbook_batch(config)
    assert result.scheme_results[0].target_total == 3000

    out = load_workbook(output_path, data_only=True)
    ws = out["sheet1"]
    assert ws["E2"].value == 1000
    assert ws["E3"].value == 2000


def test_batch_detail_export_preserves_user_named_detail_sheet(tmp_path: Path):
    input_path = tmp_path / "preserve_user_sheet.xlsx"
    output_path = tmp_path / "preserve_user_sheet_out.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "sheet1"
    ws.append(["车间", "材料", "共耗料"])
    ws.append(["一车间", 100, 0])
    user_sheet = wb.create_sheet("明细_01_共耗料分摊")
    user_sheet["A1"] = "用户自建明细"
    wb.save(input_path)

    config = BatchAllocationConfig(
        input_path=str(input_path),
        output_path=str(output_path),
        sheet_name="sheet1",
        header_row=1,
        schemes=[
            AllocationScheme("共耗料分摊", "manual", 3, [2], manual_amount=100),
        ],
    )

    allocate_workbook_batch(config)

    out = load_workbook(output_path, data_only=True)
    assert out["明细_01_共耗料分摊"]["A1"].value == "用户自建明细"
    assert "明细_01_共耗料分摊_2" in out.sheetnames
    assert out["分摊汇总"]["B6"].value == "明细_01_共耗料分摊_2"


def test_batch_allocation_rejects_duplicate_target_column(tmp_path: Path):
    input_path = tmp_path / "duplicate.xlsx"
    output_path = tmp_path / "duplicate_out.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "sheet1"
    ws.append(["车间", "材料", "费用"])
    ws.append(["一车间", 100, 0])
    wb.save(input_path)

    config = BatchAllocationConfig(
        input_path=str(input_path),
        output_path=str(output_path),
        sheet_name="sheet1",
        header_row=1,
        schemes=[
            AllocationScheme("方案1", "manual", 3, [2], manual_amount=100),
            AllocationScheme("方案2", "manual", 3, [2], manual_amount=200),
        ],
    )

    try:
        allocate_workbook_batch(config)
    except ValueError as exc:
        assert "重复使用同一分摊列" in str(exc)
    else:
        raise AssertionError("Expected duplicate target column to be rejected")


def test_preview_rejects_nonzero_amount_without_participating_base(tmp_path: Path):
    input_path = tmp_path / "zero_base.xlsx"
    output_path = tmp_path / "zero_base_out.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "sheet1"
    ws.append(["车间", "材料", "费用"])
    ws.append(["一车间", 0, 0])
    ws.append(["二车间", 0, 0])
    wb.save(input_path)

    config = BatchAllocationConfig(
        input_path=str(input_path),
        output_path=str(output_path),
        sheet_name="sheet1",
        header_row=1,
        schemes=[
            AllocationScheme("方案1", "manual", 3, [2], manual_amount=100),
        ],
    )

    try:
        preview_workbook_batch(config)
    except ValueError as exc:
        assert "没有可参与分摊的行" in str(exc)
    else:
        raise AssertionError("Expected preview to reject nonzero allocation with zero base")


def test_invalid_regex_is_rejected_before_allocation(tmp_path: Path):
    input_path = tmp_path / "invalid_regex.xlsx"
    output_path = tmp_path / "invalid_regex_out.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "sheet1"
    ws.append(["车间", "材料", "费用"])
    ws.append(["一车间", 100, 0])
    wb.save(input_path)

    config = BatchAllocationConfig(
        input_path=str(input_path),
        output_path=str(output_path),
        sheet_name="sheet1",
        header_row=1,
        schemes=[
            AllocationScheme(
                "方案1",
                "manual",
                3,
                [2],
                manual_amount=100,
                filter_rules=[FilterRule(column=1, operator="regex", value="[")],
            ),
        ],
    )

    try:
        preview_workbook_batch(config)
    except ValueError as exc:
        assert "正则表达式无效" in str(exc)
    else:
        raise AssertionError("Expected invalid regex to be rejected")


def test_create_sample_workbook(tmp_path: Path):
    path = tmp_path / "sample.xlsx"
    create_sample_workbook(str(path))

    wb = load_workbook(path, data_only=True)
    ws = wb["测试数据"]
    assert ws["A1"].value == "日期"
    assert ws["B4"].value == "销售配货部"
    assert ws.max_column == 8


def test_scheme_template_round_trip_by_current_headers(tmp_path: Path):
    path = tmp_path / "template.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "sheet1"
    ws.append(["车间", "材料", "人工", "待分摊运费", "运费分摊"])
    ws.append(["一车间", 100, 100, 1000, 0])
    wb.save(path)

    headers = get_headers(str(path), "sheet1", 1)
    schemes = [
        {
            "name": "运费分摊",
            "amount_mode": "source_column",
            "amount_column": headers[3].label,
            "manual_amount": "",
            "allocation_column": headers[4].label,
            "base_columns": [headers[1].label, headers[2].label],
            "filter_logic": "OR",
            "filter_rules": [
                {"column": headers[0].label, "operator": "equals", "value": "销售配货部"}
            ],
        }
    ]

    template = serialize_scheme_template(schemes, headers, "sheet1", 1)
    imported = import_scheme_template(template, headers)

    assert imported == schemes


def test_scheme_template_rejects_missing_columns(tmp_path: Path):
    path = tmp_path / "template_missing.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "sheet1"
    ws.append(["车间", "材料", "运费分摊"])
    ws.append(["一车间", 100, 0])
    wb.save(path)

    headers = get_headers(str(path), "sheet1", 1)
    template = {
        "version": 1,
        "template_type": "allocation_schemes",
        "schemes": [
            {
                "name": "运费分摊",
                "amount_mode": "source_column",
                "amount_column": {"label": "D列 - 待分摊运费", "header": "待分摊运费"},
                "manual_amount": "",
                "allocation_column": {"label": "E列 - 运费分摊", "header": "运费分摊"},
                "base_columns": [
                    {"label": "B列 - 材料", "header": "材料"},
                    {"label": "C列 - 人工", "header": "人工"},
                ],
                "filter_logic": "OR",
                "filter_rules": [
                    {
                        "column": {"label": "A列 - 车间", "header": "车间"},
                        "operator": "equals",
                        "value": "销售配货部",
                    }
                ],
            }
        ],
    }

    try:
        import_scheme_template(template, headers)
    except ValueError as exc:
        assert "待分摊运费" in str(exc)
        assert "人工" in str(exc)
    else:
        raise AssertionError("Expected missing template columns to be rejected")


def test_scheme_template_rejects_invalid_template_values(tmp_path: Path):
    path = tmp_path / "template_invalid.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "sheet1"
    ws.append(["车间", "材料", "费用"])
    ws.append(["一车间", 100, 0])
    wb.save(path)

    headers = get_headers(str(path), "sheet1", 1)
    wrong_type = {"template_type": "other", "schemes": []}
    try:
        import_scheme_template(wrong_type, headers)
    except ValueError as exc:
        assert "模板类型不匹配" in str(exc)
    else:
        raise AssertionError("Expected wrong template type to be rejected")

    invalid_values = {
        "template_type": "allocation_schemes",
        "schemes": [
            {
                "name": "方案1",
                "amount_mode": "bad_mode",
                "allocation_column": {"label": headers[2].label, "header": headers[2].header},
                "base_columns": [{"label": headers[1].label, "header": headers[1].header}],
                "filter_logic": "XOR",
                "filter_rules": [
                    {
                        "column": {"label": headers[0].label, "header": headers[0].header},
                        "operator": "bad_operator",
                        "value": "销售配货部",
                    }
                ],
            }
        ],
    }

    try:
        import_scheme_template(invalid_values, headers)
    except ValueError as exc:
        assert "金额来源类型无效" in str(exc)
        assert "过滤规则关系无效" in str(exc)
        assert "过滤条件类型无效" in str(exc)
    else:
        raise AssertionError("Expected invalid template values to be rejected")
