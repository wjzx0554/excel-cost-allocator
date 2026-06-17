from dataclasses import dataclass
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path
import re
from typing import Dict, Iterable, List, Optional, Sequence, Set, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


MONEY_QUANT = Decimal("0.01")


@dataclass
class ColumnInfo:
    index: int
    letter: str
    header: str

    @property
    def label(self) -> str:
        title = self.header or "空标题"
        return f"{self.letter}列 - {title}"


@dataclass
class FilterRule:
    column: int
    operator: str
    value: str = ""


@dataclass
class AllocationConfig:
    input_path: str
    output_path: str
    sheet_name: str
    header_row: int
    base_columns: Sequence[int]
    allocation_columns: Sequence[int]
    filter_column: Optional[int] = None
    excluded_values: Optional[Set[str]] = None
    filter_rules: Optional[Sequence[FilterRule]] = None
    filter_logic: str = "OR"
    detail_sheet_name: str = "分摊明细"


@dataclass
class RowAllocation:
    row_number: int
    participates: bool
    reason: str
    filter_value: str
    base_values: List[Decimal]
    base_total: Decimal
    effective_base: Decimal
    ratio: Decimal
    allocations: Dict[int, Decimal]


@dataclass
class AllocationResult:
    output_path: str
    total_rows: int
    participating_rows: int
    excluded_rows: int
    base_total: Decimal
    allocation_totals: Dict[int, Decimal]


@dataclass
class AllocationScheme:
    name: str
    amount_source: str
    allocation_column: int
    base_columns: Sequence[int]
    filter_rules: Optional[Sequence[FilterRule]] = None
    filter_logic: str = "OR"
    amount_column: Optional[int] = None
    manual_amount: Optional[Decimal] = None


@dataclass
class BatchAllocationConfig:
    input_path: str
    output_path: str
    sheet_name: str
    header_row: int
    schemes: Sequence[AllocationScheme]
    detail_sheet_name: str = "分摊汇总"


@dataclass
class SchemeAllocationResult:
    name: str
    allocation_column: int
    amount_source: str
    amount_column: Optional[int]
    manual_amount: Decimal
    total_rows: int
    participating_rows: int
    excluded_rows: int
    base_total: Decimal
    target_total: Decimal
    distributed_total: Decimal


@dataclass
class BatchAllocationResult:
    output_path: str
    total_rows: int
    scheme_results: List[SchemeAllocationResult]


@dataclass
class SchemeRun:
    scheme: AllocationScheme
    rows: List[RowAllocation]
    base_total: Decimal
    target_total: Decimal
    distributed_total: Decimal


def normalize_value(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def display_value(value: str) -> str:
    return "<空白>" if value == "" else value


def parse_number(value) -> Decimal:
    if value is None or value == "":
        return Decimal("0")
    if isinstance(value, bool):
        return Decimal("0")
    if isinstance(value, Decimal):
        return value
    if isinstance(value, int):
        return Decimal(value)
    if isinstance(value, float):
        return Decimal(str(value))
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return Decimal("0")
        text = text.replace(",", "").replace("，", "").replace("￥", "").replace("¥", "")
        try:
            return Decimal(text)
        except InvalidOperation:
            return Decimal("0")
    return Decimal("0")


def round_money(value: Decimal) -> Decimal:
    return value.quantize(MONEY_QUANT, rounding=ROUND_HALF_UP)


def get_sheet_names(file_path: str) -> List[str]:
    wb = load_workbook(file_path, read_only=False, data_only=False)
    try:
        return list(wb.sheetnames)
    finally:
        wb.close()


def guess_header_row(file_path: str, sheet_name: str, max_scan_rows: int = 20) -> int:
    wb = load_workbook(file_path, read_only=False, data_only=True)
    try:
        ws = wb[sheet_name]
        best_row = 1
        best_count = -1
        scan_to = min(ws.max_row, max_scan_rows)
        for row in range(1, scan_to + 1):
            count = _row_value_count(ws, row)
            if count > best_count:
                best_count = count
                best_row = row
        return best_row
    finally:
        wb.close()


def get_headers(file_path: str, sheet_name: str, header_row: int) -> List[ColumnInfo]:
    wb = load_workbook(file_path, read_only=False, data_only=True)
    try:
        ws = wb[sheet_name]
        max_column = _detect_used_max_column(ws, header_row)
        headers: List[ColumnInfo] = []
        for col in range(1, max_column + 1):
            value = normalize_value(ws.cell(header_row, col).value)
            headers.append(ColumnInfo(col, get_column_letter(col), value))
        return headers
    finally:
        wb.close()


def _row_value_count(ws, row_number: int) -> int:
    count = 0
    for (_row, _col), cell in ws._cells.items():
        if _row == row_number and normalize_value(cell.value):
            count += 1
    return count


def _detect_used_max_column(ws, header_row: int, scan_rows: int = 2000) -> int:
    last_col = 0
    max_scan_row = min(ws.max_row or header_row, header_row + scan_rows)

    for (_row, _col), cell in ws._cells.items():
        if header_row <= _row <= max_scan_row and normalize_value(cell.value):
            last_col = max(last_col, _col)

    for merged_range in ws.merged_cells.ranges:
        if merged_range.min_row <= header_row <= merged_range.max_row:
            last_col = max(last_col, merged_range.max_col)

    return max(last_col, 1)


def get_unique_values(
    file_path: str,
    sheet_name: str,
    header_row: int,
    column_index: int,
    max_values: int = 5000,
) -> List[str]:
    wb = load_workbook(file_path, read_only=False, data_only=True)
    try:
        ws = wb[sheet_name]
        seen = set()
        values: List[str] = []
        for row in range(header_row + 1, ws.max_row + 1):
            text = normalize_value(ws.cell(row, column_index).value)
            if text not in seen:
                seen.add(text)
                values.append(text)
                if len(values) >= max_values:
                    break
        return values
    finally:
        wb.close()


def preview_filter_matches(
    file_path: str,
    sheet_name: str,
    header_row: int,
    filter_rules: Sequence[FilterRule],
    filter_logic: str = "OR",
    sample_limit: int = 20,
) -> Tuple[int, List[Tuple[int, str]]]:
    wb = load_workbook(file_path, read_only=False, data_only=True)
    try:
        ws = wb[sheet_name]
        matches = 0
        samples: List[Tuple[int, str]] = []
        for row in range(header_row + 1, ws.max_row + 1):
            matched, reason = evaluate_filter_rules(ws, row, filter_rules, filter_logic)
            if matched:
                matches += 1
                if len(samples) < sample_limit:
                    samples.append((row, reason))
        return matches, samples
    finally:
        wb.close()


def evaluate_filter_rules(ws, row_number: int, rules: Sequence[FilterRule], logic: str = "OR") -> Tuple[bool, str]:
    active_rules = [rule for rule in rules if rule.column and rule.operator]
    if not active_rules:
        return False, ""

    results = []
    reasons = []
    for rule in active_rules:
        cell_value = normalize_value(ws.cell(row_number, rule.column).value)
        matched = _match_filter_rule(cell_value, rule)
        results.append(matched)
        if matched:
            reasons.append(_filter_reason(rule, cell_value))

    normalized_logic = (logic or "OR").upper()
    if normalized_logic == "AND":
        matched = all(results)
    else:
        matched = any(results)

    return matched, "；".join(reasons) if matched else ""


def build_legacy_filter_rules(filter_column: Optional[int], excluded_values: Optional[Set[str]]) -> List[FilterRule]:
    if not filter_column:
        return []
    return [
        FilterRule(column=filter_column, operator="equals", value=normalize_value(value))
        for value in sorted(excluded_values or set(), key=normalize_value)
    ]


def _match_filter_rule(cell_value: str, rule: FilterRule) -> bool:
    value = normalize_value(rule.value)
    operator = rule.operator

    if operator == "equals":
        return cell_value == value
    if operator == "not_equals":
        return cell_value != value
    if operator == "contains":
        return value in cell_value
    if operator == "not_contains":
        return value not in cell_value
    if operator == "regex":
        try:
            return re.search(value, cell_value) is not None
        except re.error:
            return False
    if operator == "blank":
        return cell_value == ""
    if operator == "not_blank":
        return cell_value != ""
    return False


def _filter_reason(rule: FilterRule, cell_value: str) -> str:
    letter = get_column_letter(rule.column)
    value = display_value(normalize_value(rule.value))
    actual = display_value(cell_value)
    labels = {
        "equals": "等于",
        "not_equals": "不等于",
        "contains": "包含",
        "not_contains": "不包含",
        "regex": "正则",
        "blank": "为空",
        "not_blank": "非空",
    }
    label = labels.get(rule.operator, rule.operator)
    if rule.operator in {"blank", "not_blank"}:
        return f"{letter}列 {label}，实际值：{actual}"
    return f"{letter}列 {label} {value}，实际值：{actual}"


def allocate_workbook(config: AllocationConfig) -> AllocationResult:
    _validate_config(config)

    input_path = Path(config.input_path)
    keep_vba = input_path.suffix.lower() == ".xlsm"
    formulas_wb = load_workbook(config.input_path, data_only=False, keep_vba=keep_vba)
    values_wb = load_workbook(config.input_path, data_only=True, keep_vba=keep_vba)

    try:
        ws = formulas_wb[config.sheet_name]
        values_ws = values_wb[config.sheet_name]
        data_start = config.header_row + 1
        max_row = values_ws.max_row
        filter_rules = list(config.filter_rules or [])
        if not filter_rules:
            filter_rules = build_legacy_filter_rules(config.filter_column, config.excluded_values)

        rows: List[RowAllocation] = []
        base_total = Decimal("0")

        for row_number in range(data_start, max_row + 1):
            base_values = [
                parse_number(values_ws.cell(row_number, col).value)
                for col in config.base_columns
            ]
            row_base_total = sum(base_values, Decimal("0"))
            filter_matched, filter_reason = evaluate_filter_rules(
                values_ws,
                row_number,
                filter_rules,
                config.filter_logic,
            )
            filter_value = filter_reason

            if filter_matched:
                participates = False
                reason = filter_reason or "过滤排除"
            elif row_base_total <= 0:
                participates = False
                reason = "基数小于等于0"
            else:
                participates = True
                reason = ""

            effective_base = row_base_total if participates else Decimal("0")
            base_total += effective_base
            rows.append(
                RowAllocation(
                    row_number=row_number,
                    participates=participates,
                    reason=reason,
                    filter_value=filter_value,
                    base_values=base_values,
                    base_total=row_base_total,
                    effective_base=effective_base,
                    ratio=Decimal("0"),
                    allocations={},
                )
            )

        allocation_totals: Dict[int, Decimal] = {}
        for target_col in config.allocation_columns:
            target_total = round_money(
                sum(
                    parse_number(values_ws.cell(row, target_col).value)
                    for row in range(data_start, max_row + 1)
                )
            )
            allocation_totals[target_col] = target_total
            if target_total != 0 and base_total <= 0:
                raise ValueError("没有可参与分摊的行，无法分配非零费用。")

            rounded_sum = Decimal("0")
            for item in rows:
                item.ratio = item.effective_base / base_total if base_total > 0 else Decimal("0")
                amount = round_money(target_total * item.ratio) if item.participates else Decimal("0")
                item.allocations[target_col] = amount
                rounded_sum += amount

            residual = target_total - rounded_sum
            residual_row = _find_residual_row(rows)
            if residual_row is not None:
                residual_row.allocations[target_col] = round_money(
                    residual_row.allocations[target_col] + residual
                )

            for item in rows:
                cell = ws.cell(item.row_number, target_col)
                cell.value = float(item.allocations[target_col])
                cell.number_format = '#,##0.00'

        _write_detail_sheet(
            formulas_wb,
            config,
            rows,
            base_total,
            allocation_totals,
        )

        formulas_wb.save(config.output_path)
    finally:
        formulas_wb.close()
        values_wb.close()

    participating_rows = sum(1 for item in rows if item.participates)
    return AllocationResult(
        output_path=config.output_path,
        total_rows=len(rows),
        participating_rows=participating_rows,
        excluded_rows=len(rows) - participating_rows,
        base_total=round_money(base_total),
        allocation_totals=allocation_totals,
    )


def _validate_config(config: AllocationConfig) -> None:
    if not config.input_path:
        raise ValueError("请选择 Excel 文件。")
    if not config.output_path:
        raise ValueError("请选择输出文件。")
    if config.header_row < 1:
        raise ValueError("表头行必须大于等于 1。")
    if not config.base_columns:
        raise ValueError("请至少选择一个参与占比计算的列。")
    if not config.allocation_columns:
        raise ValueError("请至少选择一个需要分配的费用列。")
    overlap = set(config.base_columns) & set(config.allocation_columns)
    if overlap:
        letters = ", ".join(get_column_letter(col) for col in sorted(overlap))
        raise ValueError(f"参与占比列和需要分配列不能重复：{letters}")


def _validate_scheme(scheme: AllocationScheme) -> None:
    if not scheme.name.strip():
        raise ValueError("分摊方案名称不能为空。")
    if not scheme.allocation_column or scheme.allocation_column < 1:
        raise ValueError(f"方案 {scheme.name} 的分摊列无效。")
    if not scheme.base_columns:
        raise ValueError(f"方案 {scheme.name} 至少要选择一个占比计算列。")
    if set(scheme.base_columns) & {scheme.allocation_column}:
        letter = get_column_letter(scheme.allocation_column)
        raise ValueError(f"方案 {scheme.name} 的分摊列不能同时作为占比计算列：{letter}")

    amount_source = (scheme.amount_source or "column_total").strip()
    if amount_source not in {"column_total", "manual"}:
        raise ValueError(f"方案 {scheme.name} 的金额来源无效。")
    if amount_source == "column_total" and scheme.amount_column is not None and scheme.amount_column < 1:
        raise ValueError(f"方案 {scheme.name} 的金额来源列无效。")
    if amount_source == "manual" and scheme.manual_amount is None:
        raise ValueError(f"方案 {scheme.name} 请选择手工金额。")
    for rule in scheme.filter_rules or []:
        _validate_filter_rule(rule, scheme.name)


def _validate_filter_rule(rule: FilterRule, scheme_name: str = "") -> None:
    label = f"方案 {scheme_name} 的" if scheme_name else ""
    if not rule.column or rule.column < 1:
        raise ValueError(f"{label}过滤条件字段无效。")
    if not rule.operator:
        raise ValueError(f"{label}过滤条件类型不能为空。")
    if rule.operator not in {
        "equals",
        "not_equals",
        "contains",
        "not_contains",
        "regex",
        "blank",
        "not_blank",
    }:
        raise ValueError(f"{label}过滤条件类型无效：{rule.operator}")
    if rule.operator == "regex":
        try:
            re.compile(normalize_value(rule.value))
        except re.error as exc:
            raise ValueError(f"{label}正则表达式无效：{exc}") from exc


def _build_filter_rules_for_scheme(scheme: AllocationScheme) -> List[FilterRule]:
    rules = list(scheme.filter_rules or [])
    return rules


def _prepare_row_allocations(
    values_ws,
    data_start: int,
    max_row: int,
    base_columns: Sequence[int],
    filter_rules: Sequence[FilterRule],
    filter_logic: str,
) -> Tuple[List[RowAllocation], Decimal]:
    rows: List[RowAllocation] = []
    base_total = Decimal("0")

    for row_number in range(data_start, max_row + 1):
        base_values = [
            parse_number(values_ws.cell(row_number, col).value)
            for col in base_columns
        ]
        row_base_total = sum(base_values, Decimal("0"))
        filter_matched, filter_reason = evaluate_filter_rules(
            values_ws,
            row_number,
            filter_rules,
            filter_logic,
        )
        filter_value = filter_reason

        if filter_matched:
            participates = False
            reason = filter_reason or "过滤排除"
        elif row_base_total <= 0:
            participates = False
            reason = "基数小于等于0"
        else:
            participates = True
            reason = ""

        effective_base = row_base_total if participates else Decimal("0")
        base_total += effective_base
        rows.append(
            RowAllocation(
                row_number=row_number,
                participates=participates,
                reason=reason,
                filter_value=filter_value,
                base_values=base_values,
                base_total=row_base_total,
                effective_base=effective_base,
                ratio=Decimal("0"),
                allocations={},
            )
        )

    return rows, base_total


def _resolve_scheme_target_total(values_ws, data_start: int, max_row: int, scheme: AllocationScheme) -> Decimal:
    amount_source = (scheme.amount_source or "column_total").strip()
    if amount_source == "manual":
        return round_money(parse_number(scheme.manual_amount))

    source_column = scheme.amount_column or scheme.allocation_column
    total = sum(
        parse_number(values_ws.cell(row_number, source_column).value)
        for row_number in range(data_start, max_row + 1)
    )
    return round_money(total)


def _apply_target_allocation(
    ws,
    rows: Sequence[RowAllocation],
    target_column: int,
    target_total: Decimal,
    base_total: Decimal,
) -> Decimal:
    if target_total != 0 and base_total <= 0:
        raise ValueError("没有可参与分摊的行，无法分配非零费用。")

    rounded_sum = Decimal("0")
    for item in rows:
        item.ratio = item.effective_base / base_total if base_total > 0 else Decimal("0")
        amount = round_money(target_total * item.ratio) if item.participates else Decimal("0")
        item.allocations[target_column] = amount
        rounded_sum += amount

    residual = target_total - rounded_sum
    residual_row = _find_residual_row(rows)
    if residual_row is not None:
        residual_row.allocations[target_column] = round_money(
            residual_row.allocations[target_column] + residual
        )

    for item in rows:
        cell = ws.cell(item.row_number, target_column)
        cell.value = float(item.allocations[target_column])
        cell.number_format = '#,##0.00'

    return round_money(sum(item.allocations[target_column] for item in rows))


def allocate_workbook_batch(config: BatchAllocationConfig) -> BatchAllocationResult:
    _validate_batch_config(config)

    input_path = Path(config.input_path)
    keep_vba = input_path.suffix.lower() == ".xlsm"
    formulas_wb = load_workbook(config.input_path, data_only=False, keep_vba=keep_vba)
    values_wb = load_workbook(config.input_path, data_only=True, keep_vba=keep_vba)

    try:
        ws = formulas_wb[config.sheet_name]
        values_ws = values_wb[config.sheet_name]
        data_start = config.header_row + 1
        max_row = values_ws.max_row
        used_target_columns = set()
        scheme_runs: List[SchemeRun] = []

        for scheme in config.schemes:
            _validate_scheme(scheme)
            if scheme.allocation_column in used_target_columns:
                raise ValueError(
                    f"多个方案不能重复使用同一分摊列：{get_column_letter(scheme.allocation_column)}"
                )
            used_target_columns.add(scheme.allocation_column)

            filter_rules = _build_filter_rules_for_scheme(scheme)
            rows, base_total = _prepare_row_allocations(
                values_ws,
                data_start,
                max_row,
                scheme.base_columns,
                filter_rules,
                scheme.filter_logic,
            )
            target_total = _resolve_scheme_target_total(values_ws, data_start, max_row, scheme)
            distributed_total = _apply_target_allocation(
                ws,
                rows,
                scheme.allocation_column,
                target_total,
                base_total,
            )
            participating_rows = sum(1 for item in rows if item.participates)
            scheme_runs.append(
                SchemeRun(
                    scheme=scheme,
                    rows=rows,
                    base_total=base_total,
                    target_total=target_total,
                    distributed_total=distributed_total,
                )
            )

        _write_batch_detail_sheet(formulas_wb, config, scheme_runs, max_row - data_start + 1)
        formulas_wb.save(config.output_path)
    finally:
        formulas_wb.close()
        values_wb.close()

    scheme_results: List[SchemeAllocationResult] = []
    for run in scheme_runs:
        participating_rows = sum(1 for item in run.rows if item.participates)
        scheme_results.append(
            SchemeAllocationResult(
                name=run.scheme.name,
                allocation_column=run.scheme.allocation_column,
                amount_source=run.scheme.amount_source,
                amount_column=run.scheme.amount_column,
                manual_amount=round_money(parse_number(run.scheme.manual_amount)),
                total_rows=len(run.rows),
                participating_rows=participating_rows,
                excluded_rows=len(run.rows) - participating_rows,
                base_total=round_money(run.base_total),
                target_total=round_money(run.target_total),
                distributed_total=round_money(run.distributed_total),
            )
        )

    total_rows = max((len(run.rows) for run in scheme_runs), default=0)
    return BatchAllocationResult(
        output_path=config.output_path,
        total_rows=total_rows,
        scheme_results=scheme_results,
    )


def preview_workbook_batch(config: BatchAllocationConfig) -> BatchAllocationResult:
    _validate_batch_config(config)

    input_path = Path(config.input_path)
    keep_vba = input_path.suffix.lower() == ".xlsm"
    values_wb = load_workbook(config.input_path, data_only=True, keep_vba=keep_vba)

    try:
        values_ws = values_wb[config.sheet_name]
        data_start = config.header_row + 1
        max_row = values_ws.max_row
        used_target_columns = set()
        scheme_results: List[SchemeAllocationResult] = []

        for scheme in config.schemes:
            _validate_scheme(scheme)
            if scheme.allocation_column in used_target_columns:
                raise ValueError(
                    f"多个方案不能重复使用同一分摊列：{get_column_letter(scheme.allocation_column)}"
                )
            used_target_columns.add(scheme.allocation_column)

            rows, base_total = _prepare_row_allocations(
                values_ws,
                data_start,
                max_row,
                scheme.base_columns,
                _build_filter_rules_for_scheme(scheme),
                scheme.filter_logic,
            )
            target_total = _resolve_scheme_target_total(values_ws, data_start, max_row, scheme)
            if target_total != 0 and base_total <= 0:
                raise ValueError(f"方案 {scheme.name} 没有可参与分摊的行，无法分配非零费用。")
            participating_rows = sum(1 for item in rows if item.participates)
            scheme_results.append(
                SchemeAllocationResult(
                    name=scheme.name,
                    allocation_column=scheme.allocation_column,
                    amount_source=scheme.amount_source,
                    amount_column=scheme.amount_column,
                    manual_amount=round_money(parse_number(scheme.manual_amount)),
                    total_rows=len(rows),
                    participating_rows=participating_rows,
                    excluded_rows=len(rows) - participating_rows,
                    base_total=round_money(base_total),
                    target_total=round_money(target_total),
                    distributed_total=round_money(target_total),
                )
            )
    finally:
        values_wb.close()

    return BatchAllocationResult(
        output_path=config.output_path,
        total_rows=max((item.total_rows for item in scheme_results), default=0),
        scheme_results=scheme_results,
    )


def create_sample_workbook(output_path: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "测试数据"
    ws.append(
        [
            "日期",
            "车间",
            "产品名称",
            "产品类别",
            "完工入库材料成本 wg_materialcost",
            "本期人工费 laborcost",
            "本期共耗料 consumptioncost",
            "运费分摊",
        ]
    )
    rows = [
        ["2026-06", "一车间", "A产品", "成品", 10000, 3000, 0, 0],
        ["2026-06", "二车间", "B产品", "成品", 20000, 5000, 0, 0],
        ["2026-06", "销售配货部", "配货费用", "内部", 5000, 1000, 0, 0],
        ["2026-06", "三车间", "C产品", "半成品", 15000, 4000, 0, 0],
    ]
    for row in rows:
        ws.append(row)
    for col in range(1, 9):
        _header(ws, 1, col, ws.cell(1, col).value)
        ws.column_dimensions[get_column_letter(col)].width = 18
    for row in range(2, 6):
        for col in range(5, 9):
            ws.cell(row, col).number_format = '#,##0.00'
    ws.freeze_panes = "A2"
    wb.save(output_path)


def _validate_batch_config(config: BatchAllocationConfig) -> None:
    if not config.input_path:
        raise ValueError("请选择 Excel 文件。")
    if not config.output_path:
        raise ValueError("请选择输出文件。")
    if config.header_row < 1:
        raise ValueError("表头行必须大于等于 1。")
    if not config.schemes:
        raise ValueError("请至少添加一个分摊方案。")


def _format_amount_source_label(scheme: AllocationScheme) -> str:
    amount_source = (scheme.amount_source or "column_total").strip()
    if amount_source == "manual":
        return f"手工金额 {round_money(parse_number(scheme.manual_amount))}"
    source_column = scheme.amount_column or scheme.allocation_column
    return f"列金额 {get_column_letter(source_column)}"


def _write_batch_detail_sheet(
    wb,
    config: BatchAllocationConfig,
    scheme_runs: Sequence[SchemeRun],
    total_rows: int,
) -> None:
    _delete_batch_detail_sheets(wb, config.detail_sheet_name)
    source_ws = wb[config.sheet_name]
    summary_name = _unique_sheet_name(config.detail_sheet_name, set(wb.sheetnames))
    summary = wb.create_sheet(summary_name)
    summary.sheet_properties.tabColor = "5B9BD5"
    summary["A1"] = "费用分摊汇总"
    summary["A1"].font = Font(size=15, bold=True, color="1F4E78")

    summary["A2"] = "原工作表"
    summary["B2"] = config.sheet_name
    summary["C2"] = "表头行"
    summary["D2"] = config.header_row
    summary["E2"] = "方案数"
    summary["F2"] = len(scheme_runs)
    summary["A3"] = "数据行数"
    summary["B3"] = total_rows
    summary["C3"] = "说明"
    summary["D3"] = "每个方案已单独生成明细工作表"

    summary_row = 5
    summary_headers = [
        "方案名称",
        "明细工作表",
        "分摊列",
        "金额来源",
        "待分摊金额",
        "参与行数",
        "不参与行数",
        "基数合计",
        "分摊后合计",
        "校验差额",
        "规则关系",
    ]
    for col, title in enumerate(summary_headers, start=1):
        _header(summary, summary_row, col, title)

    used_names = set(wb.sheetnames)
    for offset, run in enumerate(scheme_runs, start=1):
        sheet_name = _unique_sheet_name(
            _clean_sheet_name(f"明细_{offset:02d}_{run.scheme.name}"),
            used_names,
        )
        used_names.add(sheet_name)
        _write_scheme_detail_sheet(
            wb,
            source_ws,
            config,
            run,
            sheet_name,
        )

        row = summary_row + offset
        participating = sum(1 for item in run.rows if item.participates)
        amount_source = "手工输入" if (run.scheme.amount_source or "column_total") == "manual" else _format_amount_source_label(run.scheme)
        difference = round_money(run.target_total - run.distributed_total)
        summary.cell(row, 1, run.scheme.name)
        summary.cell(row, 2, sheet_name)
        summary.cell(row, 2).hyperlink = f"#{_sheet_reference(sheet_name)}!A1"
        summary.cell(row, 2).style = "Hyperlink"
        summary.cell(row, 3, get_column_letter(run.scheme.allocation_column))
        summary.cell(row, 4, amount_source)
        summary.cell(row, 5, float(round_money(run.target_total)))
        summary.cell(row, 6, participating)
        summary.cell(row, 7, len(run.rows) - participating)
        summary.cell(row, 8, float(round_money(run.base_total)))
        summary.cell(row, 9, float(round_money(run.distributed_total)))
        summary.cell(row, 10, float(difference))
        summary.cell(row, 11, run.scheme.filter_logic or "OR")
        for col in (5, 8, 9, 10):
            summary.cell(row, col).number_format = '#,##0.00'

    summary.freeze_panes = summary.cell(summary_row + 1, 1).coordinate
    summary.auto_filter.ref = f"A{summary_row}:K{summary_row + len(scheme_runs)}"
    widths = [24, 24, 10, 18, 14, 10, 12, 14, 14, 12, 10]
    for col, width in enumerate(widths, start=1):
        summary.column_dimensions[get_column_letter(col)].width = width


def _write_scheme_detail_sheet(
    wb,
    source_ws,
    config: BatchAllocationConfig,
    run: SchemeRun,
    sheet_name: str,
) -> None:
    scheme = run.scheme
    detail = wb.create_sheet(sheet_name)
    detail.sheet_properties.tabColor = "70AD47"
    detail["A1"] = f"方案明细：{scheme.name}"
    detail["A1"].font = Font(size=15, bold=True, color="375623")

    participating = sum(1 for item in run.rows if item.participates)
    detail["A2"] = "原工作表"
    detail["B2"] = config.sheet_name
    detail["C2"] = "分摊列"
    detail["D2"] = get_column_letter(scheme.allocation_column)
    detail["E2"] = "金额来源"
    detail["F2"] = _format_amount_source_label(scheme)
    detail["A3"] = "待分摊金额"
    detail["B3"] = float(round_money(run.target_total))
    detail["C3"] = "参与行数"
    detail["D3"] = participating
    detail["E3"] = "不参与行数"
    detail["F3"] = len(run.rows) - participating
    detail["A4"] = "基数合计"
    detail["B4"] = float(round_money(run.base_total))
    detail["C4"] = "分摊后合计"
    detail["D4"] = float(round_money(run.distributed_total))
    detail["E4"] = "校验差额"
    detail["F4"] = float(round_money(run.target_total - run.distributed_total))
    detail["A5"] = "规则关系"
    detail["B5"] = scheme.filter_logic or "OR"
    for cell in ("B3", "B4", "D4", "F4"):
        detail[cell].number_format = '#,##0.00'

    source_columns = set(scheme.base_columns) | {scheme.allocation_column}
    if scheme.amount_column:
        source_columns.add(scheme.amount_column)
    for rule in scheme.filter_rules or []:
        source_columns.add(rule.column)
    source_headers = {
        col: normalize_value(source_ws.cell(config.header_row, col).value) or get_column_letter(col)
        for col in source_columns
    }

    filter_summary_row = 7
    detail.cell(filter_summary_row, 1, "过滤条件")
    detail.cell(filter_summary_row, 1).font = Font(bold=True, color="1F4E78")
    if scheme.filter_rules:
        filter_headers = ["字段", "条件", "值"]
        for col, title in enumerate(filter_headers, start=1):
            _header(detail, filter_summary_row + 1, col, title)
        for offset, rule in enumerate(scheme.filter_rules, start=1):
            row = filter_summary_row + 1 + offset
            detail.cell(row, 1, f"{get_column_letter(rule.column)}列 - {source_headers.get(rule.column, '')}")
            detail.cell(row, 2, _operator_label(rule.operator))
            detail.cell(row, 3, display_value(normalize_value(rule.value)))
        table_row = filter_summary_row + len(scheme.filter_rules) + 4
    else:
        detail.cell(filter_summary_row + 1, 1, "无过滤条件")
        table_row = filter_summary_row + 4

    headers = ["源行号", "是否参与", "不参与原因", "过滤命中说明"]
    headers.extend(
        f"基数列 {get_column_letter(col)} - {source_headers[col]}"
        for col in scheme.base_columns
    )
    headers.extend(
        [
            "分摊基数",
            "占比",
            f"分摊结果 {get_column_letter(scheme.allocation_column)} - {source_headers[scheme.allocation_column]}",
        ]
    )
    for col, title in enumerate(headers, start=1):
        _header(detail, table_row, col, title)

    for row_offset, item in enumerate(run.rows, start=1):
        row = table_row + row_offset
        col = 1
        detail.cell(row, col, item.row_number)
        col += 1
        detail.cell(row, col, "是" if item.participates else "否")
        col += 1
        detail.cell(row, col, item.reason)
        col += 1
        detail.cell(row, col, item.filter_value)
        col += 1
        for value in item.base_values:
            detail.cell(row, col, float(value))
            detail.cell(row, col).number_format = '#,##0.00'
            col += 1
        detail.cell(row, col, float(item.effective_base))
        detail.cell(row, col).number_format = '#,##0.00'
        col += 1
        detail.cell(row, col, float(item.ratio))
        detail.cell(row, col).number_format = '0.0000%'
        col += 1
        detail.cell(row, col, float(item.allocations.get(scheme.allocation_column, Decimal("0"))))
        detail.cell(row, col).number_format = '#,##0.00'

    detail.freeze_panes = detail.cell(table_row + 1, 1).coordinate
    detail.auto_filter.ref = f"A{table_row}:{get_column_letter(len(headers))}{table_row + len(run.rows)}"
    for col in range(1, len(headers) + 1):
        detail.column_dimensions[get_column_letter(col)].width = 16
    detail.column_dimensions["A"].width = 10
    detail.column_dimensions["B"].width = 10
    detail.column_dimensions["C"].width = 18
    detail.column_dimensions["D"].width = 22


def _delete_batch_detail_sheets(wb, summary_sheet_name: str) -> None:
    for name in list(wb.sheetnames):
        sheet = wb[name]
        marker = normalize_value(sheet["A1"].value)
        is_summary_name = name == summary_sheet_name or name.startswith(f"{summary_sheet_name}_")
        if is_summary_name and marker == "费用分摊汇总":
            del wb[name]
        elif name == "分摊明细" and marker == "费用分摊明细":
            del wb[name]
        elif re.match(r"^明细_\d{2}_", name) and marker.startswith("方案明细："):
            del wb[name]


def _clean_sheet_name(name: str) -> str:
    cleaned = re.sub(r"[\[\]\:\*\?\/\\]", "_", name or "明细")
    cleaned = cleaned.strip("' ") or "明细"
    return cleaned[:31]


def _unique_sheet_name(base_name: str, used_names: Set[str]) -> str:
    base_name = _clean_sheet_name(base_name)
    if base_name not in used_names:
        return base_name
    for index in range(2, 1000):
        suffix = f"_{index}"
        candidate = f"{base_name[:31 - len(suffix)]}{suffix}"
        if candidate not in used_names:
            return candidate
    raise ValueError("明细工作表名称过多，无法自动命名。")


def _sheet_reference(sheet_name: str) -> str:
    return "'" + sheet_name.replace("'", "''") + "'"


def _operator_label(operator: str) -> str:
    labels = {
        "equals": "等于",
        "not_equals": "不等于",
        "contains": "包含",
        "not_contains": "不包含",
        "regex": "正则",
        "blank": "为空",
        "not_blank": "非空",
    }
    return labels.get(operator, operator)


def _find_residual_row(rows: Iterable[RowAllocation]) -> Optional[RowAllocation]:
    participating = [item for item in rows if item.participates and item.effective_base > 0]
    if not participating:
        return None
    return max(participating, key=lambda item: (item.effective_base, -item.row_number))


def _header(ws, row: int, col: int, value: str) -> None:
    cell = ws.cell(row, col, value)
    cell.fill = PatternFill("solid", fgColor="305496")
    cell.font = Font(bold=True, color="FFFFFF")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = Border(
        left=Side(style="thin", color="BFBFBF"),
        right=Side(style="thin", color="BFBFBF"),
        top=Side(style="thin", color="BFBFBF"),
        bottom=Side(style="thin", color="BFBFBF"),
    )


def _write_detail_sheet(
    wb,
    config: AllocationConfig,
    rows: Sequence[RowAllocation],
    base_total: Decimal,
    allocation_totals: Dict[int, Decimal],
) -> None:
    if config.detail_sheet_name in wb.sheetnames:
        del wb[config.detail_sheet_name]

    detail = wb.create_sheet(config.detail_sheet_name)
    detail.sheet_properties.tabColor = "5B9BD5"
    detail["A1"] = "费用分摊明细"
    detail["A1"].font = Font(size=15, bold=True, color="1F4E78")

    participating = sum(1 for item in rows if item.participates)
    detail["A2"] = "原工作表"
    detail["B2"] = config.sheet_name
    detail["C2"] = "表头行"
    detail["D2"] = config.header_row
    detail["A3"] = "数据行数"
    detail["B3"] = len(rows)
    detail["C3"] = "参与行数"
    detail["D3"] = participating
    detail["E3"] = "不参与行数"
    detail["F3"] = len(rows) - participating
    detail["A4"] = "有效分摊基数合计"
    detail["B4"] = float(round_money(base_total))
    detail["B4"].number_format = '#,##0.00'

    source_ws = wb[config.sheet_name]
    source_headers = {
        col: normalize_value(source_ws.cell(config.header_row, col).value) or get_column_letter(col)
        for col in set(config.base_columns) | set(config.allocation_columns)
    }

    row_cursor = 6
    summary_headers = ["费用列", "原始合计", "分摊后合计"]
    for col, title in enumerate(summary_headers, start=1):
        _header(detail, row_cursor, col, title)
    for offset, col_index in enumerate(config.allocation_columns, start=1):
        row = row_cursor + offset
        total = allocation_totals[col_index]
        distributed = sum(item.allocations[col_index] for item in rows)
        detail.cell(row, 1, f"{get_column_letter(col_index)}列 - {source_headers[col_index]}")
        detail.cell(row, 2, float(total))
        detail.cell(row, 3, float(distributed))
        detail.cell(row, 2).number_format = '#,##0.00'
        detail.cell(row, 3).number_format = '#,##0.00'

    table_row = row_cursor + len(config.allocation_columns) + 3
    headers = ["源行号", "是否参与", "不参与原因", "过滤列值"]
    headers.extend(
        f"基数列 {get_column_letter(col)} - {source_headers[col]}"
        for col in config.base_columns
    )
    headers.extend(["分摊基数", "占比"])
    headers.extend(
        f"分摊结果 {get_column_letter(col)} - {source_headers[col]}"
        for col in config.allocation_columns
    )

    for col, title in enumerate(headers, start=1):
        _header(detail, table_row, col, title)

    for row_offset, item in enumerate(rows, start=1):
        row = table_row + row_offset
        col = 1
        detail.cell(row, col, item.row_number)
        col += 1
        detail.cell(row, col, "是" if item.participates else "否")
        col += 1
        detail.cell(row, col, item.reason)
        col += 1
        detail.cell(row, col, item.filter_value)
        col += 1
        for value in item.base_values:
            detail.cell(row, col, float(value))
            detail.cell(row, col).number_format = '#,##0.00'
            col += 1
        detail.cell(row, col, float(item.effective_base))
        detail.cell(row, col).number_format = '#,##0.00'
        col += 1
        detail.cell(row, col, float(item.ratio))
        detail.cell(row, col).number_format = '0.0000%'
        col += 1
        for target_col in config.allocation_columns:
            detail.cell(row, col, float(item.allocations[target_col]))
            detail.cell(row, col).number_format = '#,##0.00'
            col += 1

    detail.freeze_panes = detail.cell(table_row + 1, 1).coordinate
    detail.auto_filter.ref = f"A{table_row}:{get_column_letter(len(headers))}{table_row + len(rows)}"

    for col in range(1, len(headers) + 1):
        detail.column_dimensions[get_column_letter(col)].width = 16
    detail.column_dimensions["A"].width = 10
    detail.column_dimensions["C"].width = 16
    detail.column_dimensions["D"].width = 18
